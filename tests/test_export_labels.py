from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from docx import Document
from openpyxl import load_workbook

from app.core.compare import CompareResult, STATUT_MODIFIE
from app.core.report import write_report
from app.core.word_report import write_ftm_document
from app.extract.pdf_reader import PdfExtraction


class ExportLabelTests(unittest.TestCase):
    def test_excel_uses_market_and_after_ftm_headers(self) -> None:
        table = pd.DataFrame([{
            "niveau": "R+2",
            "occupation": "VASCULAIRE ANGIO",
            "piece": "Consultation 1",
            "numero": "28",
            "room_id": "room-1",
            "source_room": "Vasculaire 01",
            "source_material": "PC 10/16A 2P+T",
            "origin": "pdf",
            "categorie": "Électricité",
            "materiel": "Prise de courant",
            "quantite_avant": 8,
            "quantite_apres": 6,
            "ecart": -2,
            "statut": STATUT_MODIFIE,
            "pages": "5",
            "labels": "glyphe",
            "rapprochement": "correspondance objet",
        }])
        result = CompareResult(table=table, niveau="R+2")

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "comparatif.xlsx"
            write_report(str(path), result, PdfExtraction(), "marché.xlsx", "ftm.pdf")
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                headers = [cell.value for cell in next(workbook["Comparatif"].iter_rows())]
            finally:
                workbook.close()

        self.assertIn("Quantité marché", headers)
        self.assertIn("Quantité après FTM", headers)
        self.assertNotIn("Qté avant (maquette)", headers)
        self.assertNotIn("Qté après (plan)", headers)

    def test_word_uses_new_material_table_headers_and_computes_prix_total(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "ftm.docx"
            write_ftm_document(path, {
                "materials": [{
                    "id": "row-1",
                    "room": "Vasculaire 01",
                    "material": "PC 10/16A 2P+T",
                    "quantity_before": "8",
                    "quantity_after": "6",
                    "unit_price": "10",
                }],
            })
            document = Document(path)

        material_table = next(
            table for table in document.tables
            if " ".join(table.rows[0].cells[0].text.split()) == "Nom de la pièce"
        )
        top_header = [" ".join(cell.text.split()) for cell in material_table.rows[0].cells]
        bottom_header = [" ".join(cell.text.split()) for cell in material_table.rows[1].cells]
        self.assertIn("Lot", top_header)
        self.assertIn("Sous lot", top_header)
        self.assertIn("Prestations", top_header)
        self.assertIn("Quantité marché", top_header)
        self.assertIn("Quantité FTM", top_header)
        self.assertIn("À compléter par l'émetteur", top_header)
        self.assertIn("Prix unitaire", bottom_header)
        self.assertIn("Prix total", bottom_header)

        data_row = [" ".join(cell.text.split()) for cell in material_table.rows[3].cells]
        self.assertEqual(data_row[3], "PC 10/16A 2P+T")
        self.assertEqual(data_row[7], "60,00 €")

        subtotal_row = [" ".join(cell.text.split()) for cell in material_table.rows[4].cells]
        self.assertIn("Lot sous-total", subtotal_row)
        self.assertIn("60,00 € H.T.", subtotal_row)


if __name__ == "__main__":
    unittest.main()
