# -*- coding: utf-8 -*-
"""Validation structurelle d'un classeur FTMgen genere."""
import argparse
import unicodedata

from openpyxl import load_workbook


def _without_accents(value: str) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFD", value)
        if unicodedata.category(char) != "Mn"
    )


def validate(path: str) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=True)
    try:
        required = {
            "Synthese",
            "Comparatif",
            "Ecarts uniquement",
            "A valider",
            "Tracabilite plan",
        }
        actual = {_without_accents(name) for name in workbook.sheetnames}
        missing = required - actual
        if missing:
            raise ValueError(f"Onglets absents: {sorted(missing)}")

        rows = workbook["Comparatif"].max_row - 1
        if rows < 1:
            raise ValueError("La feuille Comparatif est vide")
        return rows, len(workbook.sheetnames)
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    args = parser.parse_args()
    rows, sheets = validate(args.workbook)
    print(f"VALIDO: {args.workbook} | {rows} lineas | {sheets} hojas")


if __name__ == "__main__":
    main()
