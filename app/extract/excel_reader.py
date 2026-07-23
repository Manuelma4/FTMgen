# -*- coding: utf-8 -*-
"""Lecture du listing maquette "Pieces + Materiel" (export type MEDIVIE).

Structures acceptees :
- une feuille avec une colonne Niveau ;
- une feuille par niveau, avec le niveau porte par le nom de la feuille.

Chaque feuille contient une ligne d'en-tete puis une ligne par materiel. Les
colonnes piece/numero/niveau peuvent n'etre renseignees que sur la premiere
ligne de chaque piece (cellules fusionnees) -> forward-fill.
"""
import re
import unicodedata

import openpyxl
import pandas as pd

from ..core.relations import excel_room_id, excel_room_label

COLUMNS = ["occupation", "piece", "numero", "niveau", "categorie", "code_article", "materiel", "quantite"]
HEADER_FIRST_CELL = "occupation"

HEADER_ALIASES = {
    "occupation": "occupation",
    "nom de la piece": "piece",
    "piece": "piece",
    "numero": "numero",
    "no": "numero",
    "n lot": "numero",
    "lot": "numero",
    "niveau": "niveau",
    "categorie": "categorie",
    "code article": "code_article",
    "code": "code_article",
    "materiel": "materiel",
    "article": "materiel",
    "quantite": "quantite",
    "qte": "quantite",
}
REQUIRED_BASE_COLUMNS = {"occupation", "piece", "categorie", "materiel", "quantite"}
IGNORED_SHEETS = {"notation", "notations", "mode d emploi", "correspondance pieces", "correspondance articles"}


def _norm(value) -> str:
    value = unicodedata.normalize("NFD", str(value or ""))
    value = "".join(c for c in value if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def read_listing(path: str) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        frames = []
        errors = []
        for sheet in _candidate_sheets(workbook):
            rows = [row for row in sheet.iter_rows(values_only=True)]
            try:
                frame = _read_sheet_rows(rows, sheet.title)
            except ValueError as exc:
                errors.append(f"{sheet.title}: {exc}")
                continue
            if not frame.empty:
                frames.append(frame)
    finally:
        workbook.close()

    if not frames:
        detail = "; ".join(errors[:4])
        suffix = f" Detail : {detail}" if detail else ""
        raise ValueError(
            "Feuille Pieces + Materiel introuvable : aucune feuille ne contient "
            "les colonnes obligatoires Occupation, Nom de la piece, Categorie, "
            f"Materiel et Quantite.{suffix}"
        )

    return pd.concat(frames, ignore_index=True).reset_index(drop=True)


def _read_sheet_rows(rows: list[tuple], sheet_name: str) -> pd.DataFrame:
    header_idx = None
    for index, row in enumerate(rows):
        labels = {_norm(value) for value in row if value is not None}
        if HEADER_FIRST_CELL in labels:
            header_idx = index
            break
    if header_idx is None:
        raise ValueError("aucune ligne d'en-tete contenant 'Occupation'")

    header = rows[header_idx]
    positions = {}
    for index, label in enumerate(header):
        canonical = HEADER_ALIASES.get(_norm(label))
        if canonical and canonical not in positions:
            positions[canonical] = index

    missing = REQUIRED_BASE_COLUMNS - set(positions)
    if missing:
        raise ValueError(f"colonnes obligatoires absentes : {sorted(missing)}")

    data = []
    for row in rows[header_idx + 1:]:
        data.append({
            column: row[positions[column]] if column in positions and positions[column] < len(row) else None
            for column in COLUMNS
        })
    frame = pd.DataFrame(data, columns=COLUMNS).dropna(how="all")

    for column in ("occupation", "piece", "numero", "niveau"):
        if column in positions:
            frame[column] = frame[column].ffill()

    level = _sheet_level(sheet_name, rows)
    if "niveau" not in positions or frame["niveau"].dropna().astype(str).str.strip().eq("").all():
        frame["niveau"] = level
    else:
        frame["niveau"] = frame["niveau"].fillna(level)

    frame = frame.dropna(subset=["materiel"])
    for column in ("occupation", "piece", "niveau", "categorie", "code_article", "materiel"):
        frame[column] = frame[column].fillna("").astype(str).str.strip()
    frame["numero"] = frame["numero"].apply(lambda value: str(value).strip() if value is not None else "")
    frame["quantite"] = pd.to_numeric(frame["quantite"], errors="coerce").fillna(0).astype(int)
    frame = frame[frame["materiel"] != ""]

    # Un nom de pièce n'est pas une identité physique : un même niveau peut
    # contenir plusieurs « Consultation » appartenant à des occupations/lots
    # différents.  Ces colonnes suivent chaque ligne matériel afin que le
    # comparatif ne soit jamais obligé de regrouper par le seul libellé.
    frame["room_id"] = frame.apply(
        lambda row: excel_room_id(
            row.get("niveau"), row.get("occupation"), row.get("piece"), row.get("numero")
        ),
        axis=1,
    )
    frame["room_label"] = frame.apply(
        lambda row: excel_room_label(
            row.get("niveau"), row.get("occupation"), row.get("piece"), row.get("numero")
        ),
        axis=1,
    )

    placeholder = frame["materiel"].str.match(r"^\(.*\)$")
    frame.loc[placeholder, "materiel"] = ""
    frame.loc[placeholder, "quantite"] = 0
    return frame.reset_index(drop=True)


def read_correspondences(path: str) -> tuple[dict, dict]:
    """Lit les feuilles optionnelles de correspondance pieces et articles."""
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    room_map, material_map = {}, {}
    try:
        for sheet_name in workbook.sheetnames:
            normalized = _norm(sheet_name)
            if "correspondance" not in normalized:
                continue
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_norm(value) for value in rows[0]]
            if "piece plan pdf" in headers and "piece existante excel" in headers:
                left, right = headers.index("piece plan pdf"), headers.index("piece existante excel")
                for row in rows[1:]:
                    if len(row) > max(left, right) and row[left] and row[right]:
                        room_map[str(row[left]).strip()] = str(row[right]).strip()
            if "article plan pdf" in headers and "materiel existant excel" in headers:
                left, right = headers.index("article plan pdf"), headers.index("materiel existant excel")
                for row in rows[1:]:
                    if len(row) > max(left, right) and row[left] and row[right]:
                        material_map[str(row[left]).strip()] = str(row[right]).strip()
    finally:
        workbook.close()
    return room_map, material_map


def _candidate_sheets(workbook):
    for name in workbook.sheetnames:
        normalized = _norm(name)
        if normalized in IGNORED_SHEETS or "correspondance" in normalized:
            continue
        yield workbook[name]


def _sheet_level(sheet_name: str, rows: list[tuple]) -> str:
    if str(sheet_name or "").strip():
        return str(sheet_name).strip()
    for row in rows[:5]:
        for value in row:
            text = str(value or "").strip()
            if text:
                return text
    return str(sheet_name).strip()
